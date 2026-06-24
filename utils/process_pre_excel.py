#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
预处理Excel数据处理工具
处理Data目录下的total.xlsx文件，将状态为Pre的数据进行字段转移
"""

import pandas as pd
import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def process_total_excel_data(input_file: str = "Data/total.xlsx", output_file: str = None) -> pd.DataFrame:
    """
    处理total.xlsx文件，将状态为Pre的数据从h字段转移到新字段

    Args:
        input_file: 输入Excel文件路径
        output_file: 输出文件路径，如果为None则覆盖原文件

    Returns:
        处理后的DataFrame
    """
    logger.info(f"开始处理Excel文件: {input_file}")

    # 检查文件是否存在
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"输入文件不存在: {input_file}")

    # 读取Excel数据
    try:
        df = pd.read_excel(input_file)
        logger.info(f"成功读取Excel数据，共 {len(df)} 行")
    except Exception as e:
        logger.error(f"读取Excel文件失败: {e}")
        raise

    # 检查必要的列是否存在
    required_columns = ['Status', 'RS1h', 'RS2h', 'RS3h', 'RS4h']
    missing_columns = [col for col in required_columns if col not in df.columns]

    if missing_columns:
        raise ValueError(f"缺少必要的列: {missing_columns}")

    # 统计处理前的数据
    pre_count = len(df[df['Status'] == 'Pre'])
    logger.info(f"找到 {pre_count} 行状态为'Pre'的数据")

    if pre_count == 0:
        logger.warning("没有找到状态为'Pre'的数据，无需处理")
        return df

    # 创建新的字段
    # 检查是否已存在h0字段，如果不存在则创建
    for i in range(1, 5):
        h0_col = f'RS{i}h0'
        if h0_col not in df.columns:
            df[h0_col] = None
            logger.info(f"创建新字段: {h0_col}")

    # 处理状态为Pre的数据
    processed_count = 0

    for idx, row in df.iterrows():
        if row['Status'] == 'Pre':
            # 记录处理前的值
            h_values = {
                'RS1h': row['RS1h'],
                'RS2h': row['RS2h'],
                'RS3h': row['RS3h'],
                'RS4h': row['RS4h']
            }

            # 将h字段的数据转移到h0字段
            if not pd.isna(row['RS1h']):
                df.at[idx, 'RS1h0'] = row['RS1h']
            if not pd.isna(row['RS2h']):
                df.at[idx, 'RS2h0'] = row['RS2h']
            if not pd.isna(row['RS3h']):
                df.at[idx, 'RS3h0'] = row['RS3h']
            if not pd.isna(row['RS4h']):
                df.at[idx, 'RS4h0'] = row['RS4h']

            # 清空h字段的数据
            df.at[idx, 'RS1h'] = None
            df.at[idx, 'RS2h'] = None
            df.at[idx, 'RS3h'] = None
            df.at[idx, 'RS4h'] = None

            processed_count += 1

            # 记录处理详情
            logger.debug(f"处理第 {idx} 行: {h_values} -> h0字段")

    logger.info(f"成功处理 {processed_count} 行数据")

    # 数据质量检查
    for rs_field in ['RS1', 'RS2', 'RS3', 'RS4']:
        h0_not_null = df[f'{rs_field}h0'].notna().sum()
        h_null = df[f'{rs_field}h'].isna().sum()
        logger.info(f"{rs_field}: h0非空数据 {h0_not_null} 个, h空数据 {h_null} 个")

    # 保存处理后的数据
    if output_file is None:
        output_file = input_file
        logger.info(f"将覆盖原文件: {output_file}")
    else:
        logger.info(f"将保存到新文件: {output_file}")

    # 确保输出目录存在
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"创建输出目录: {output_dir}")

    # 使用常规方法保存Excel文件
    try:
        df.to_excel(output_file, index=False)
        logger.info(f"Excel文件已保存到: {output_file}")
    except Exception as e:
        logger.error(f"保存Excel文件失败: {e}")
        raise

    return df

def save_excel_with_charts(input_file: str, output_file: str, df: pd.DataFrame):
    """
    保存Excel文件，保留原有的图表和工作表格式

    Args:
        input_file: 原始Excel文件路径
        output_file: 输出Excel文件路径
        df: 要保存的DataFrame
    """
    try:
        # 加载原始工作簿以保留图表和格式
        wb = load_workbook(input_file)

        # 获取第一个工作表（假设数据在第一个工作表）
        ws = wb.active

        # 清除现有数据（保留图表）
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

        # 将DataFrame写入工作表，从第二行开始（保留表头）
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 保存工作簿
        wb.save(output_file)
        logger.info(f"Excel文件已保存到: {output_file}")

    except Exception as e:
        # 如果图表保存失败，使用常规方法保存
        logger.warning(f"图表保存失败，使用常规保存方法: {e}")
        df.to_excel(output_file, index=False)
        logger.info(f"Excel文件已保存到: {output_file}")

def backup_original_excel_file(input_file: str) -> str:
    """
    备份原始Excel文件

    Args:
        input_file: 原始文件路径

    Returns:
        备份文件路径
    """
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"文件不存在: {input_file}")

    # 创建备份文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(input_file)[0]
    extension = os.path.splitext(input_file)[1]
    backup_file = f"{base_name}_backup_{timestamp}{extension}"

    # 复制文件
    import shutil
    shutil.copy2(input_file, backup_file)
    logger.info(f"已创建备份文件: {backup_file}")

    return backup_file

def validate_excel_data_integrity(df: pd.DataFrame) -> bool:
    """
    验证Excel数据完整性

    Args:
        df: 处理后的DataFrame

    Returns:
        是否通过验证
    """
    logger.info("开始Excel数据完整性验证")

    # 检查数据行数
    if len(df) == 0:
        logger.error("数据为空")
        return False

    # 检查关键字段的数据类型
    numeric_fields = ['RS1h0', 'RS2h0', 'RS3h0', 'RS4h0', 'RS1h', 'RS2h', 'RS3h', 'RS4h']
    for field in numeric_fields:
        if field in df.columns:
            # 检查是否有非数值数据
            non_numeric = df[field].apply(lambda x: not pd.isna(x) and not isinstance(x, (int, float, complex)))
            if non_numeric.any():
                logger.warning(f"{field} 字段包含非数值数据")

    # 检查状态字段
    if 'Status' in df.columns:
        status_counts = df['Status'].value_counts()
        logger.info(f"状态分布: {status_counts.to_dict()}")

    logger.info("Excel数据完整性验证完成")
    return True

def main():
    """主函数"""
    import sys

    # 解析命令行参数
    input_file = "Data/total.xlsx"
    output_file = None
    backup = True

    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    if len(sys.argv) > 3:
        backup = sys.argv[3].lower() not in ['false', 'no', '0']

    print("=" * 60)
    print("预处理Excel数据处理工具")
    print("=" * 60)

    try:
        # 备份原始文件
        if backup:
            backup_file = backup_original_excel_file(input_file)

        # 处理数据
        processed_df = process_total_excel_data(input_file, output_file)

        # 验证数据完整性
        if validate_excel_data_integrity(processed_df):
            print("[SUCCESS] 数据处理成功完成！")
        else:
            print("[WARNING] 数据处理完成，但完整性验证有警告")

        print(f"[INFO] 处理统计:")
        print(f"   - 总数据行数: {len(processed_df)}")

        if 'Status' in processed_df.columns:
            pre_status_count = (processed_df['Status'] == 'Pre').sum()
            print(f"   - 处理前Pre状态数据: {pre_status_count} 行")

        # 统计字段数据
        for rs_field in ['RS1', 'RS2', 'RS3', 'RS4']:
            h0_count = processed_df[f'{rs_field}h0'].notna().sum()
            h_count = processed_df[f'{rs_field}h'].notna().sum()
            print(f"   - {rs_field}: h0有数据 {h0_count} 个, h有数据 {h_count} 个")

        print(f"[OUTPUT] 输出文件: {output_file if output_file else input_file}")
        if backup:
            print(f"[BACKUP] 备份文件: {backup_file}")

        # 验证数据转移效果
        print(f"\n[VALIDATION] 数据转移验证:")
        pre_data = processed_df[processed_df['Status'] == 'Pre'].iloc[:3]
        for rs_field in ['RS1', 'RS2', 'RS3', 'RS4']:
            h0_val = pre_data[f'{rs_field}h0'].iloc[0]
            h_val = pre_data[f'{rs_field}h'].iloc[0]
            h0_status = "有数据" if not pd.isna(h0_val) else "空"
            h_status = "有数据" if not pd.isna(h_val) else "空"
            print(f"   - {rs_field}: h0={h0_status}({h0_val}), h={h_status}({h_val})")

    except Exception as e:
        logger.error(f"处理失败: {e}")
        print(f"[ERROR] 处理失败: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()