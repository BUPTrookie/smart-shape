#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel工作表数据处理工具（改进版）
处理Excel文件中的指定工作表，将状态为Pre的数据进行字段转移
"""

import pandas as pd
import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def clear_worksheet_data(ws):
    """
    清除工作表中的数据，但保留格式和图表

    Args:
        ws: openpyxl工作表对象
    """
    # 获取工作表的最大行列数
    max_row = ws.max_row
    max_col = ws.max_column

    # 清除所有单元格的值，但保留格式
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None

    logger.info(f"已清除工作表 {ws.title} 的数据（{max_row}行 x {max_col}列）")

def process_excel_sheets(input_file: str, output_file: str = None) -> None:
    """
    处理Excel文件中的指定工作表

    Args:
        input_file: 输入Excel文件路径
        output_file: 输出文件路径，如果为None则覆盖原文件
    """

    # 需要处理的4个工作表（匹配实际工作表名称）
    target_sheets = [
        "Reshaping (含重复)",  # 注意空格和括号的精确匹配
        "Reshaping （去重复）",
        "Reshaping（去重复去跳点）",
        "Reshaping (NG)"
    ]

    logger.info(f"开始处理Excel文件: {input_file}")

    # 检查文件是否存在
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"输入文件不存在: {input_file}")

    # 读取Excel文件的所有工作表名称
    try:
        xl_file = pd.ExcelFile(input_file)
        available_sheets = xl_file.sheet_names
        logger.info(f"Excel文件包含工作表: {available_sheets}")
    except Exception as e:
        logger.error(f"读取Excel文件失败: {e}")
        raise

    # 精确匹配目标工作表
    matched_sheets = []
    for target in target_sheets:
        if target in available_sheets:
            matched_sheets.append(target)
        else:
            logger.warning(f"工作表 '{target}' 不存在")

    if not matched_sheets:
        raise ValueError("没有找到可处理的工作表")

    logger.info(f"匹配到的工作表: {matched_sheets}")

    # 加载工作簿以保留所有格式和图表
    try:
        wb = load_workbook(input_file)
        logger.info("成功加载Excel工作簿")
    except Exception as e:
        logger.error(f"加载工作簿失败: {e}")
        raise

    total_processed = 0

    # 处理每个目标工作表
    for sheet_name in matched_sheets:
        logger.info(f"正在处理工作表: {sheet_name}")

        # 读取工作表数据
        try:
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            logger.info(f"工作表 {sheet_name} 包含 {len(df)} 行数据，{len(df.columns)} 列")
        except Exception as e:
            logger.error(f"读取工作表 {sheet_name} 失败: {e}")
            continue

        # 检查必要的列是否存在
        required_columns = ['Status', 'RS1h1', 'RS2h1', 'RS3h1', 'RS4h1', 'RS1h0', 'RS2h0', 'RS3h0', 'RS4h0']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            logger.warning(f"工作表 {sheet_name} 缺少必要的列: {missing_columns}")
            logger.info(f"工作表 {sheet_name} 实际列名: {list(df.columns)}")
            continue

        # 统计处理前的数据
        pre_count = len(df[df['Status'] == 'Pre'])
        logger.info(f"工作表 {sheet_name} 找到 {pre_count} 行状态为'Pre'的数据")

        if pre_count == 0:
            logger.info(f"工作表 {sheet_name} 没有Pre状态数据，跳过处理")
            # 仍然需要清除并写回原始数据
            processed_df = df
        else:
            # 处理状态为Pre的数据
            sheet_processed_count = 0

            # 创建df的副本以避免修改原数据
            processed_df = df.copy()

            for idx, row in df.iterrows():
                if row['Status'] == 'Pre':
                    # 将h1的数据转移到h0
                    processed_df.at[idx, 'RS1h0'] = row['RS1h1']
                    processed_df.at[idx, 'RS2h0'] = row['RS2h1']
                    processed_df.at[idx, 'RS3h0'] = row['RS3h1']
                    processed_df.at[idx, 'RS4h0'] = row['RS4h1']

                    # 清空h1的数据
                    processed_df.at[idx, 'RS1h1'] = None
                    processed_df.at[idx, 'RS2h1'] = None
                    processed_df.at[idx, 'RS3h1'] = None
                    processed_df.at[idx, 'RS4h1'] = None

                    sheet_processed_count += 1

            logger.info(f"工作表 {sheet_name} 成功处理 {sheet_processed_count} 行数据")
            total_processed += sheet_processed_count

            # 验证处理结果
            h0_non_null = processed_df[processed_df['Status'] == 'Pre']['RS1h0'].notna().sum()
            h1_null = processed_df[processed_df['Status'] == 'Pre']['RS1h1'].isna().sum()
            logger.info(f"工作表 {sheet_name} 验证: h0非空 {h0_non_null} 个, h1空 {h1_null} 个")

        # 将处理后的数据写回工作表
        try:
            ws = wb[sheet_name]

            # 清除现有数据（保留格式和图表）
            clear_worksheet_data(ws)

            # 将DataFrame写入工作表
            for r_idx, row in enumerate(dataframe_to_rows(processed_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            logger.info(f"工作表 {sheet_name} 数据已更新完成")

        except Exception as e:
            logger.error(f"更新工作表 {sheet_name} 失败: {e}")
            continue

    logger.info(f"总共处理了 {total_processed} 行Pre状态数据")

    # 保存处理后的文件
    if output_file is None:
        output_file = input_file
        logger.info(f"将覆盖原文件: {output_file}")
    else:
        logger.info(f"将保存到新文件: {output_file}")

    # 确保输出目录存在
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"创建输出目录: {output_dir}")

    # 保存工作簿
    try:
        wb.save(output_file)
        logger.info(f"Excel文件已保存到: {output_file}")
    except Exception as e:
        logger.error(f"保存Excel文件失败: {e}")
        raise

def backup_original_excel_file(input_file: str) -> str:
    """
    备份原始Excel文件

    Args:
        input_file: 原始文件路径

    Returns:
        备份文件路径
    """
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"文件不存在: {input_file}")

    # 创建备份文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(input_file)[0]
    extension = os.path.splitext(input_file)[1]
    backup_file = f"{base_name}_backup_{timestamp}{extension}"

    # 复制文件
    import shutil
    shutil.copy2(input_file, backup_file)
    logger.info(f"已创建备份文件: {backup_file}")

    return backup_file

def main():
    """主函数"""
    import sys

    # 解析命令行参数
    input_file = "Data/total.xlsx"
    output_file = None
    backup = True

    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    if len(sys.argv) > 3:
        backup = sys.argv[3].lower() not in ['false', 'no', '0']

    print("=" * 60)
    print("Excel工作表数据处理工具（改进版）")
    print("=" * 60)

    try:
        # 备份原始文件
        if backup:
            backup_file = backup_original_excel_file(input_file)

        # 处理数据
        process_excel_sheets(input_file, output_file)

        print("[SUCCESS] Excel文件处理成功完成！")
        print(f"[OUTPUT] 输出文件: {output_file if output_file else input_file}")
        if backup:
            print(f"[BACKUP] 备份文件: {backup_file}")

    except Exception as e:
        logger.error(f"处理失败: {e}")
        print(f"[ERROR] 处理失败: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()