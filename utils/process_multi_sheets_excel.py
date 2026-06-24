#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel多工作表预处理工具
处理Data目录下的total.xlsx文件，处理指定的4个工作表，将状态为Pre的数据进行字段转移
"""

import pandas as pd
import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from copy import deepcopy

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def process_excel_sheets(input_file: str, output_file: str = None) -> None:
    """
    处理Excel文件中的指定工作表

    Args:
        input_file: 输入Excel文件路径
        output_file: 输出文件路径，如果为None则覆盖原文件
    """

    # 需要处理的4个工作表
    target_sheets = [
        "Reshaping(含重复)",
        "Reshaping （去重复）",
        "Reshaping（去重复去跳点）",
        "Reshaping (NG)"
    ]

    logger.info(f"开始处理Excel文件: {input_file}")

    # 检查文件是否存在
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"输入文件不存在: {input_file}")

    # 读取Excel文件的所有工作表名称
    try:
        xl_file = pd.ExcelFile(input_file)
        available_sheets = xl_file.sheet_names
        logger.info(f"Excel文件包含工作表: {available_sheets}")
    except Exception as e:
        logger.error(f"读取Excel文件失败: {e}")
        raise

    # 检查目标工作表是否存在
    missing_sheets = [sheet for sheet in target_sheets if sheet not in available_sheets]
    if missing_sheets:
        logger.warning(f"以下工作表不存在: {missing_sheets}")
        # 尝试匹配可用的工作表（处理编码问题）
        available_target_sheets = []
        for target_sheet in target_sheets:
            for available_sheet in available_sheets:
                # 简化匹配：检查关键字
                if "含重复" in target_sheet and "含重复" in available_sheet:
                    available_target_sheets.append(available_sheet)
                    break
                elif "去重复" in target_sheet and "去重复" in available_sheet and "跳点" not in available_sheet:
                    available_target_sheets.append(available_sheet)
                    break
                elif "去重复去跳点" in target_sheet and "去重复" in available_sheet and "跳点" in available_sheet:
                    available_target_sheets.append(available_sheet)
                    break
                elif "NG" in target_sheet and "NG" in available_sheet:
                    available_target_sheets.append(available_sheet)
                    break
        target_sheets = available_target_sheets
        logger.info(f"匹配到的工作表: {target_sheets}")

    if not target_sheets:
        raise ValueError("没有找到可处理的工作表")

    # 加载工作簿以保留所有格式和图表
    try:
        wb = load_workbook(input_file)
        logger.info("成功加载Excel工作簿")
    except Exception as e:
        logger.error(f"加载工作簿失败: {e}")
        raise

    # 处理每个目标工作表
    processed_count = 0

    for sheet_name in target_sheets:
        logger.info(f"正在处理工作表: {sheet_name}")

        # 检查工作表是否存在
        if sheet_name not in wb.sheetnames:
            logger.warning(f"工作表 {sheet_name} 不存在，跳过")
            continue

        # 读取工作表数据
        try:
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            logger.info(f"工作表 {sheet_name} 包含 {len(df)} 行数据")
        except Exception as e:
            logger.error(f"读取工作表 {sheet_name} 失败: {e}")
            continue

        # 检查必要的列是否存在
        required_columns = ['Status', 'RS1h1', 'RS2h1', 'RS3h1', 'RS4h1', 'RS1h0', 'RS2h0', 'RS3h0', 'RS4h0']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            logger.warning(f"工作表 {sheet_name} 缺少必要的列: {missing_columns}")
            continue

        # 统计处理前的数据
        pre_count = len(df[df['Status'] == 'Pre'])
        logger.info(f"工作表 {sheet_name} 找到 {pre_count} 行状态为'Pre'的数据")

        if pre_count == 0:
            logger.info(f"工作表 {sheet_name} 没有Pre状态数据，跳过处理")
            continue

        # 处理状态为Pre的数据
        sheet_processed_count = 0

        for idx, row in df.iterrows():
            if row['Status'] == 'Pre':
                # 将h1的数据转移到h0
                df.at[idx, 'RS1h0'] = row['RS1h1']
                df.at[idx, 'RS2h0'] = row['RS2h1']
                df.at[idx, 'RS3h0'] = row['RS3h1']
                df.at[idx, 'RS4h0'] = row['RS4h1']

                # 清空h1的数据
                df.at[idx, 'RS1h1'] = None
                df.at[idx, 'RS2h1'] = None
                df.at[idx, 'RS3h1'] = None
                df.at[idx, 'RS4h1'] = None

                sheet_processed_count += 1

        logger.info(f"工作表 {sheet_name} 成功处理 {sheet_processed_count} 行数据")
        processed_count += sheet_processed_count

        # 将处理后的数据写回工作表
        try:
            ws = wb[sheet_name]

            # 清除现有数据（保留格式和图表）
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

            # 将DataFrame写入工作表
            from openpyxl.utils.dataframe import dataframe_to_rows
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            logger.info(f"工作表 {sheet_name} 数据已更新")

        except Exception as e:
            logger.error(f"更新工作表 {sheet_name} 失败: {e}")
            continue

    logger.info(f"总共处理了 {processed_count} 行Pre状态数据")

    # 保存处理后的文件
    if output_file is None:
        output_file = input_file
        logger.info(f"将覆盖原文件: {output_file}")
    else:
        logger.info(f"将保存到新文件: {output_file}")

    # 确保输出目录存在
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"创建输出目录: {output_dir}")

    # 保存工作簿
    try:
        wb.save(output_file)
        logger.info(f"Excel文件已保存到: {output_file}")
    except Exception as e:
        logger.error(f"保存Excel文件失败: {e}")
        raise

def backup_original_excel_file(input_file: str) -> str:
    """
    备份原始Excel文件

    Args:
        input_file: 原始文件路径

    Returns:
        备份文件路径
    """
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"文件不存在: {input_file}")

    # 创建备份文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(input_file)[0]
    extension = os.path.splitext(input_file)[1]
    backup_file = f"{base_name}_backup_{timestamp}{extension}"

    # 复制文件
    import shutil
    shutil.copy2(input_file, backup_file)
    logger.info(f"已创建备份文件: {backup_file}")

    return backup_file

def main():
    """主函数"""
    import sys

    # 解析命令行参数
    input_file = "Data/total.xlsx"
    output_file = None
    backup = True

    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    if len(sys.argv) > 3:
        backup = sys.argv[3].lower() not in ['false', 'no', '0']

    print("=" * 60)
    print("Excel多工作表预处理工具")
    print("=" * 60)

    try:
        # 备份原始文件
        if backup:
            backup_file = backup_original_excel_file(input_file)

        # 处理数据
        process_excel_sheets(input_file, output_file)

        print("[SUCCESS] Excel文件处理成功完成！")
        print(f"[OUTPUT] 输出文件: {output_file if output_file else input_file}")
        if backup:
            print(f"[BACKUP] 备份文件: {backup_file}")

    except Exception as e:
        logger.error(f"处理失败: {e}")
        print(f"[ERROR] 处理失败: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()